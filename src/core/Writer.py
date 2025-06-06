import openpyxl
from . import Network, Athlete

class Writer:

    def __init__(self, network: Network, min_matches: int, max_matches: int, file_name: str):
        self.network = network
        self.min_matches = min_matches
        self.max_matches = max_matches
        self.file_name = file_name
        self._athletes: list[Athlete] = []

    def search(self) -> None:
        self.network.setup_payload_on_search()
        self.parse_athletes()
        self.write_athletes()

    def parse_athletes(self) -> None:
        while True:
            athlete_divs = self.network.scrap_athletes_raw_data()
            if not athlete_divs:
                break
            for athlete_div in athlete_divs:
                self.process_athlete(athlete_div)

            self.network.next_page()
        self.network.reset_payload()

    def process_athlete(self, athlete_div) -> None:
        athlete = self.network.div_to_athlete(athlete_div)
        if not (self.min_matches <= athlete.total_matches() <= self.max_matches):
            return
        name_elem = athlete_div.find(class_='card-title')
        athlete.name = name_elem.text.strip()
        athlete.age = int(name_elem.find_next_sibling(class_='card-title').text.split(':')[-1])
        athlete.club = athlete_div.find('h6', string='Società').find_next('p').text.strip()
        self._athletes.append(athlete)

    def write_athletes(self) -> None:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        headers = ["Name", "Age", "Club", "Matches", "Wins", "Losses", "Draws"]

        for col_num, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_num, value=header)

        row_num = 2
        for athlete in self._athletes:
            sheet.cell(row=row_num, column=1, value=athlete.name)
            sheet.cell(row=row_num, column=2, value=athlete.age)
            sheet.cell(row=row_num, column=3, value=athlete.club)
            sheet.cell(row=row_num, column=4, value=athlete.total_matches())
            sheet.cell(row=row_num, column=5, value=athlete.wins)
            sheet.cell(row=row_num, column=6, value=athlete.losses)
            sheet.cell(row=row_num, column=7, value=athlete.draws)
            row_num += 1

        workbook.save(f"{self.file_name}.xlsx")
        workbook.close()