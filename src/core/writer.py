import openpyxl
from .service import FpiService
from .athlete import FpiAthlete


class FpiWriter:
    """Writer class responsible for exporting athlete data to Excel files."""

    def __init__(self, service: FpiService, min_matches: int, max_matches: int, file_name: str):
        """
        Initialize the writer with search parameters.

        Args:
            service: FpiService instance for data operations
            min_matches: Minimum number of matches to filter
            max_matches: Maximum number of matches to filter
            file_name: Name of the output Excel file (without extension)
        """
        self.service = service
        self.min_matches = min_matches
        self.max_matches = max_matches
        self.file_name = file_name
        self._athletes: list[FpiAthlete] = []

    def search_and_write(self) -> None:
        """
        Performs the complete search and write operation.
        Searches for athletes matching the criteria and exports them to Excel.
        """
        print("Starting athlete search...")
        self._search_athletes()

        if self._athletes:
            print(f"Found {len(self._athletes)} athletes. Writing to Excel...")
            self._write_to_excel()
            print(f"Successfully exported {len(self._athletes)} athletes to {self.file_name}.xlsx")
        else:
            print("No athletes found matching the specified criteria.")

    def _search_athletes(self) -> None:
        """Searches for athletes using the service and stores them internally."""
        try:
            self._athletes = self.service.search_athletes_with_filters(
                self.min_matches,
                self.max_matches
            )
        except Exception as e:
            print(f"Error during athlete search: {e}")
            raise

    def _write_to_excel(self) -> None:
        """Writes the found athletes to an Excel file."""
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Athletes"

            # Define headers
            headers = ["Name", "Age", "Club", "Total Matches", "Wins", "Losses", "Draws"]

            # Write headers
            for col_num, header in enumerate(headers, start=1):
                sheet.cell(row=1, column=col_num, value=header)

            # Write athlete data
            for row_num, athlete in enumerate(self._athletes, start=2):
                sheet.cell(row=row_num, column=1, value=athlete.name)
                sheet.cell(row=row_num, column=2, value=athlete.age)
                sheet.cell(row=row_num, column=3, value=athlete.club)
                sheet.cell(row=row_num, column=4, value=athlete.total_matches())
                sheet.cell(row=row_num, column=5, value=athlete.wins)
                sheet.cell(row=row_num, column=6, value=athlete.losses)
                sheet.cell(row=row_num, column=7, value=athlete.draws)

            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception as e:
                        raise e
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                sheet.column_dimensions[column_letter].width = adjusted_width

            # Save the workbook
            filename = f"{self.file_name}.xlsx"
            workbook.save(filename)
            workbook.close()

        except Exception as e:
            print(f"Error writing to Excel file: {e}")
            raise

    def get_athletes(self) -> list[FpiAthlete]:
        """
        Returns the list of athletes found in the last search.

        Returns:
            List of FpiAthlete objects
        """
        return self._athletes.copy()

    def get_athlete_count(self) -> int:
        """
        Returns the number of athletes found in the last search.

        Returns:
            Number of athletes
        """
        return len(self._athletes)

    def clear_athletes(self) -> None:
        """Clears the internal athlete list."""
        self._athletes.clear()