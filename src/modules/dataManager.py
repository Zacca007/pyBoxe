import openpyxl
from PyQt6.QtWidgets import QMessageBox
from modules import NetManager

class DataManager:
    def __init__(self, net_manager: NetManager, min_matches: int, max_matches: int, file_name: str):
        self.net_manager = net_manager
        self.min_matches = min_matches
        self.max_matches = max_matches
        self.file_name = file_name

    def search(self) -> None:
        """
        Initiates the search by fixing the payload and then fetching and displaying athletes.
        """
        self.net_manager.fix_payload()
        self.fetch_and_display()

    def fetch_and_display(self) -> None:
        """
        Fetches athlete data page by page, filters them based on match count,
        and writes the results to an Excel file.
        """
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        headers = ["Name", "Age", "Club", "Matches", "Wins", "Losses", "Draws"]

        # Scrittura degli header
        for col_num, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_num, value=header)

        row_num = 2  # Righe successive alla prima
        try:
            while True:
                athlete_divs = self.net_manager.get_athletes()
                if not athlete_divs:
                    break

                for athlete_div in athlete_divs:
                    button = athlete_div.find('button', class_='btn btn-dark btn-sm record')
                    athlete_id = button["data-id"]
                    stats = self.net_manager.get_athlete_stats(athlete_id)

                    # Filtraggio per numero di match
                    match_count = stats["matches"]
                    if match_count < self.min_matches or match_count > self.max_matches:
                        continue

                    # Estrazione dettagli atleta
                    name_elem = athlete_div.find(class_='card-title')
                    name = name_elem.text.strip() if name_elem else "Unknown"

                    age_text = name_elem.find_next_sibling(class_='card-title').text
                    age = int(age_text.split(':')[-1]) if age_text and ':' in age_text else None

                    club_elem = athlete_div.find('h6', string='Società')
                    club = club_elem.find_next('p').text.strip() if club_elem else "Unknown"

                    # Scrittura diretta nel file Excel
                    sheet.cell(row=row_num, column=1, value=name)
                    sheet.cell(row=row_num, column=2, value=age)
                    sheet.cell(row=row_num, column=3, value=club)
                    sheet.cell(row=row_num, column=4, value=match_count)
                    sheet.cell(row=row_num, column=5, value=stats["wins"])
                    sheet.cell(row=row_num, column=6, value=stats["losses"])
                    sheet.cell(row=row_num, column=7, value=stats["draws"])

                    row_num += 1

                self.net_manager.next_page()

            self.net_manager.reset_payload()

            # Salvataggio file
            workbook.save(f"{self.file_name}.xlsx")

            msg = QMessageBox()
            msg.setIcon(QMessageBox.Icon.Information)
            msg.setWindowTitle("Process Completed")
            msg.setText(f"File '{self.file_name}.xlsx' created successfully!")
            msg.setStandardButtons(QMessageBox.StandardButton.Ok)
            msg.exec()

        except Exception as e:
            QMessageBox.critical(None, "Unable to save the file", str(e))