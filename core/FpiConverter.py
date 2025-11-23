import openpyxl
from .FpiAthlete import FpiAthlete


def write_to_excel(athletes: list[FpiAthlete], filepath: str) -> None:
    """
    Scrive gli atleti trovati in un file Excel.
    
    Args:
        athletes: Lista di atleti da scrivere
        filepath: Percorso completo del file (inclusa estensione .xlsx)
    
    Raises:
        Exception: Se c'è un errore nella scrittura del file
    """
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Athletes"
        
        # Definisci gli headers
        headers = ["Nome", "Età", "Società", "Match Totali", "Vittorie", "Sconfitte", "Pareggi"]
        
        # Scrivi gli headers
        for col_num, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
        
        # Scrivi i dati degli atleti
        for row_num, athlete in enumerate(athletes, start=2):
            sheet.cell(row=row_num, column=1, value=athlete.name)
            sheet.cell(row=row_num, column=2, value=athlete.age)
            sheet.cell(row=row_num, column=3, value=athlete.club)
            sheet.cell(row=row_num, column=4, value=athlete.total_matches())
            sheet.cell(row=row_num, column=5, value=athlete.wins)
            sheet.cell(row=row_num, column=6, value=athlete.losses)
            sheet.cell(row=row_num, column=7, value=athlete.draws)
        
        # Auto-adjust larghezza colonne
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap a 50 caratteri
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Salva il workbook
        workbook.save(filepath)
        workbook.close()
        
    except Exception as e:
        print(f"Errore nella scrittura del file Excel: {e}")
        raise